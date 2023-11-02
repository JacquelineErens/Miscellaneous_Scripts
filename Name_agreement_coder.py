import os
import sys
import csv
from openpyxl import load_workbook
INPUT_FILE = sys.argv[1] #excel datafile
OUTPUT_FILE = INPUT_FILE[:-5]+"_output.csv" #csv summary file
print(OUTPUT_FILE)
def read_file(file_name): #reads the excel files - file name comes from files_list. Also has the dictionaries for the specific words you want. You'll need to change this if you want to look at more/different words
    fn = file_name
    questions_dict = {}
    percentages_dict = {}
    wb = load_workbook(filename = fn, read_only=True) #load the excel file using openxyl
    ws = wb['Sheet0'] #go to sheet0 specifically
    cols_counter = 0
    for row in ws.iter_rows(min_row=1, max_row = 1): #just get the question labels first so you can tell what images corresponded to what responses
        for cell in row:
            questions_dict[cell.value] = {}
            percentages_dict[cell.value] = {}
            cols_counter += 1
    keys_list = list(questions_dict.keys())
    for i in range(cols_counter):
        for row in ws.iter_rows(min_row=1, max_row = 1, min_col = i+1, max_col = i+1): #start at second row to skip headers, go down each column of names
            for cell in row:
                print(cell.value)
                #if cell.value in headers_list:
                    #print(cell.value)
                    #print("DUPLICATE")
                #else:
                    #headers_list.append(cell.value)
        for row in ws.iter_rows(min_row=2, min_col = i+1, max_col = i+1): #start at second row to skip headers, go down each column of names
            #print(row)
            #print("\n")
            for cell in row:
                #print(cell.value)
                if cell.value != '' and cell.value != None:
                    if cell.value.lower().strip() not in questions_dict[keys_list[i]]:
                        questions_dict[keys_list[i]].update({cell.value.lower().strip(): 1})
                    else:
                        #print("RAN ELSE")
                        questions_dict[keys_list[i]][cell.value.lower().strip()] += 1
                    if cell.value.lower().strip() not in percentages_dict[keys_list[i]]:
                        percentages_dict[keys_list[i]].update({cell.value.lower().strip(): 0})
            #print(questions_dict)
    #question dictionary set up correctly at this point
    return questions_dict, keys_list, cols_counter, percentages_dict

def get_percentages(questions_dict, keys_list, cols_counter, percentages_dict):
    for i in range(cols_counter):
        total_responses = sum(list(questions_dict[keys_list[i]].values())) #the number of people who had responses in that column
        for smaller_dict in questions_dict[keys_list[i]]:
            percentages_dict[keys_list[i]].update({smaller_dict : questions_dict[keys_list[i]][smaller_dict]/total_responses})
    return percentages_dict

def nested_dictionary_csv_writer(proportions_dictionary, numerics_dictionary, file_name, label_list = ["QuestionNumber","N_Responses","Label","Number","Proportion"]):
    #key is the file name
    #value is the dictionary with the lines the target word appeared in
    with open(file_name,"w",newline="") as f:
        thewriter = csv.writer(f)
        thewriter.writerow(label_list)
        for item in numerics_dictionary.items():
            n_responses = sum(list(item[1].values()))
            thewriter.writerow([item[0], sum(list(item[1].values()))]) #item[0] is the question number, item[1] is the dictionary of responses and their frequencies
            inner_dictionary = item[1]
            for response_and_freq in inner_dictionary.items():
                response = response_and_freq[0]
                freq = response_and_freq[1]
                prop = proportions_dictionary[item[0]][response]
                thewriter.writerow(['','',response,freq,prop])#If your data structure is a nested csv, use this one instead


    #return self.word_dict
def main():
    questions_dict, keys_list, cols_counter, percentages_dict = read_file(INPUT_FILE)
    percentages_dict = get_percentages(questions_dict, keys_list, cols_counter, percentages_dict)
    nested_dictionary_csv_writer(percentages_dict, questions_dict, file_name = OUTPUT_FILE)
main()
